import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _read_file(path):
    """根據副檔名自動選擇讀取方式，回傳 DataFrame（或 dict of DataFrame）"""
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return pd.read_excel(path, sheet_name=None)
    elif ext == '.csv':
        return pd.read_csv(path)
    else:
        raise ValueError(f'不支援的檔案格式：{ext}（僅支援 .xlsx / .csv）')


def load_invoice(path):
    """讀取發票檔案（XLSX 或 CSV），回傳整理後的 DataFrame"""
    raw = _read_file(path)

    if isinstance(raw, dict):
        # XLSX：優先找「發票資料」sheet，找不到就用第一個 sheet
        if '發票資料' in raw:
            df = raw['發票資料']
        else:
            df = list(raw.values())[0]
    else:
        df = raw

    df['賣場編號'] = df['明細備註'].astype(str).str.strip()
    df['總計'] = pd.to_numeric(df['總計'], errors='coerce').fillna(0)
    return df[['發票號碼', '發票日期', '發票狀態', '總計', '賣場編號']].copy()


def load_trade(path):
    """讀取 8591 交易檔案（XLSX 或 CSV），自動偵測並跳過備註行"""
    raw = _read_file(path)

    if isinstance(raw, dict):
        df = list(raw.values())[0]
    else:
        df = raw

    # 自動偵測：如果第一行的賣場編號不是 S 開頭，視為備註行跳過
    first_val = str(df.iloc[0]['賣場編號']).strip() if len(df) > 0 else ''
    if not first_val.startswith('S'):
        df = df.iloc[1:].reset_index(drop=True)
    df['賣場編號'] = df['賣場編號'].astype(str).str.strip()
    df['金額'] = pd.to_numeric(df['金額'], errors='coerce').fillna(0)
    df['手續費'] = pd.to_numeric(df['手續費'], errors='coerce').fillna(0)
    df['交易所得'] = pd.to_numeric(df['交易所得'], errors='coerce').fillna(0)
    return df[['賣場編號', '購買時間', '完成時間', '遊戲名', '品項', '金額', '手續費', '交易所得']].copy()


def reconcile(xlsx_df, csv_df):
    """核心對帳邏輯，回傳對帳結果 dict"""
    # 分離 8591 交易紀錄與銀行紀錄
    mask_8591 = xlsx_df['賣場編號'].str.startswith('S')
    xlsx_8591 = xlsx_df[mask_8591].copy()
    excluded_xlsx = xlsx_df[~mask_8591].copy()

    xlsx_ids = set(xlsx_8591['賣場編號'])
    csv_ids = set(csv_df['賣場編號'])
    both = xlsx_ids & csv_ids

    matched = []
    amount_mismatch = []

    for sid in sorted(both):
        x_row = xlsx_8591[xlsx_8591['賣場編號'] == sid].iloc[0]
        c_row = csv_df[csv_df['賣場編號'] == sid].iloc[0]
        record = {
            '賣場編號': sid,
            '發票號碼': x_row['發票號碼'],
            '發票金額': x_row['總計'],
            '交易金額': c_row['金額'],
            '遊戲名': c_row['遊戲名'],
            '品項': c_row['品項'],
            '手續費': c_row['手續費'],
            '交易所得': c_row['交易所得'],
            '購買時間': c_row['購買時間'],
            '完成時間': c_row['完成時間'],
        }
        if abs(x_row['總計'] - c_row['金額']) < 0.01:
            matched.append(record)
        else:
            record['差額'] = x_row['總計'] - c_row['金額']
            amount_mismatch.append(record)

    only_in_xlsx = []
    for sid in sorted(xlsx_ids - csv_ids):
        row = xlsx_8591[xlsx_8591['賣場編號'] == sid].iloc[0]
        only_in_xlsx.append({
            '賣場編號': sid,
            '發票號碼': row['發票號碼'],
            '發票金額': row['總計'],
        })

    only_in_csv = []
    for sid in sorted(csv_ids - xlsx_ids):
        row = csv_df[csv_df['賣場編號'] == sid].iloc[0]
        only_in_csv.append({
            '賣場編號': sid,
            '交易金額': row['金額'],
            '遊戲名': row['遊戲名'],
            '手續費': row['手續費'],
            '交易所得': row['交易所得'],
        })

    excluded_list = []
    for _, row in excluded_xlsx.iterrows():
        excluded_list.append({
            '發票號碼': row['發票號碼'],
            '賣場編號': row['賣場編號'],
            '發票金額': row['總計'],
        })

    xlsx_8591_total = xlsx_8591['總計'].sum()
    csv_total = csv_df['金額'].sum()

    return {
        'matched': matched,
        'amount_mismatch': amount_mismatch,
        'only_in_xlsx': only_in_xlsx,
        'only_in_csv': only_in_csv,
        'excluded_xlsx': excluded_list,
        'xlsx_8591_total': xlsx_8591_total,
        'csv_total': csv_total,
        'xlsx_full_total': xlsx_df['總計'].sum(),
    }


def get_summary_text(result):
    """產生摘要文字"""
    lines = []
    n_match = len(result['matched'])
    n_mismatch = len(result['amount_mismatch'])
    n_only_xlsx = len(result['only_in_xlsx'])
    n_only_csv = len(result['only_in_csv'])
    n_excluded = len(result['excluded_xlsx'])

    lines.append(f"吻合筆數：{n_match}")
    lines.append(f"金額不符：{n_mismatch}")
    lines.append(f"只在發票（無交易）：{n_only_xlsx}")
    lines.append(f"只在交易（無發票）：{n_only_csv}")
    lines.append(f"排除（銀行紀錄）：{n_excluded}")
    lines.append("")
    lines.append(f"發票 8591 總額：{result['xlsx_8591_total']:,.0f}")
    lines.append(f"交易平台總額：  {result['csv_total']:,.0f}")
    diff = result['xlsx_8591_total'] - result['csv_total']
    lines.append(f"差額：          {diff:,.0f}")

    if n_mismatch == 0 and n_only_xlsx == 0 and n_only_csv == 0:
        lines.append("\n對帳結果：全部吻合")
    else:
        lines.append(f"\n對帳結果：有 {n_mismatch + n_only_xlsx + n_only_csv} 筆差異需確認")

    return '\n'.join(lines)


def export_report(result, output_path):
    """將對帳結果匯出成 Excel 報表"""
    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def write_sheet(ws, title, headers, rows):
        ws.title = title
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_data in rows:
            ws.append([row_data.get(h, '') for h in headers])
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet1: 對帳摘要
    ws1 = wb.active
    ws1.title = '對帳摘要'
    summary_items = [
        ('吻合筆數', len(result['matched'])),
        ('金額不符', len(result['amount_mismatch'])),
        ('只在發票（無交易）', len(result['only_in_xlsx'])),
        ('只在交易（無發票）', len(result['only_in_csv'])),
        ('排除（銀行紀錄）', len(result['excluded_xlsx'])),
        ('', ''),
        ('發票 8591 總額', result['xlsx_8591_total']),
        ('交易平台總額', result['csv_total']),
        ('差額', result['xlsx_8591_total'] - result['csv_total']),
        ('', ''),
        ('發票全部總額（含銀行）', result['xlsx_full_total']),
    ]
    ws1.append(['項目', '數值'])
    for col_idx in range(1, 3):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    for item, val in summary_items:
        ws1.append([item, val])
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 18

    # Sheet2: 吻合明細
    ws2 = wb.create_sheet()
    match_headers = ['賣場編號', '發票號碼', '發票金額', '交易金額', '遊戲名', '品項', '手續費', '交易所得', '購買時間', '完成時間']
    write_sheet(ws2, '吻合明細', match_headers, result['matched'])

    # Sheet3: 差異明細
    ws3 = wb.create_sheet()
    ws3.title = '差異明細'
    # 金額不符
    row_idx = 1
    ws3.cell(row=row_idx, column=1, value='【金額不符】').font = Font(bold=True, size=12)
    row_idx += 1
    if result['amount_mismatch']:
        mismatch_headers = ['賣場編號', '發票號碼', '發票金額', '交易金額', '差額', '遊戲名']
        for col_idx, h in enumerate(mismatch_headers, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row_idx += 1
        for r in result['amount_mismatch']:
            for col_idx, h in enumerate(mismatch_headers, 1):
                ws3.cell(row=row_idx, column=col_idx, value=r.get(h, ''))
            row_idx += 1
    else:
        ws3.cell(row=row_idx, column=1, value='（無）')
        row_idx += 1

    row_idx += 1
    ws3.cell(row=row_idx, column=1, value='【只在發票（無對應交易）】').font = Font(bold=True, size=12)
    row_idx += 1
    if result['only_in_xlsx']:
        ox_headers = ['賣場編號', '發票號碼', '發票金額']
        for col_idx, h in enumerate(ox_headers, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row_idx += 1
        for r in result['only_in_xlsx']:
            for col_idx, h in enumerate(ox_headers, 1):
                ws3.cell(row=row_idx, column=col_idx, value=r.get(h, ''))
            row_idx += 1
    else:
        ws3.cell(row=row_idx, column=1, value='（無）')
        row_idx += 1

    row_idx += 1
    ws3.cell(row=row_idx, column=1, value='【只在交易（無對應發票）】').font = Font(bold=True, size=12)
    row_idx += 1
    if result['only_in_csv']:
        oc_headers = ['賣場編號', '交易金額', '遊戲名', '手續費', '交易所得']
        for col_idx, h in enumerate(oc_headers, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row_idx += 1
        for r in result['only_in_csv']:
            for col_idx, h in enumerate(oc_headers, 1):
                ws3.cell(row=row_idx, column=col_idx, value=r.get(h, ''))
            row_idx += 1
    else:
        ws3.cell(row=row_idx, column=1, value='（無）')
        row_idx += 1

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet4: 排除紀錄
    ws4 = wb.create_sheet()
    ex_headers = ['發票號碼', '賣場編號', '發票金額']
    write_sheet(ws4, '排除紀錄', ex_headers, result['excluded_xlsx'])

    wb.save(output_path)
